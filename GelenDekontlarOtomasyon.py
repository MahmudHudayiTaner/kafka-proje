"""
Mail'den okunmamış ve konusu 'Dekont' olan mailler çekilir ve bir excele kaydedilir.

"""
from openpyxl import load_workbook
import pandas as pd
import imaplib
import email
from email.header import decode_header

from pdfminer.high_level import extract_text
import os
import json
import google.generativeai as genai

import time

dosya_yolu = "config.xlsx"

# Config Excelinin okunması ve Config Sözlüğün oluşturulması
def excel_to_dict(excel_path):
    df = pd.read_excel(excel_path, engine='openpyxl')

    return dict(zip(df['Name'], df['Value']))

config = excel_to_dict(dosya_yolu)

EMAIL = config["mail"]
PASSWORD = config["mail_app_password"]
IMAP_SERVER = config["imap_server"]  # Gmail için: imap.gmail.com
EXCEL_YOLU = config["dekont_excel"]

API_KEY = config["gemini_api_key"]



# ***********************************************************************************

def extract_pdfminer(pdf_path):
    print("=== pdfminer.six ile çıkarılıyor ===")
    text = extract_text(pdf_path)
    return text


def append_pdf_data_to_excel(bilgiler, dosya_yolu):
    from openpyxl import load_workbook

    wb = load_workbook(dosya_yolu)
    ws = wb.active
    satir = ws.max_row + 1

    ws.cell(row=satir, column=1, value=bilgiler.get("İşlem Tarihi", ""))
    ws.cell(row=satir, column=2, value=bilgiler.get("Gönderen Kişi", ""))
    ws.cell(row=satir, column=3, value=bilgiler.get("Sorgu Numarası", ""))
    ws.cell(row=satir, column=4, value=bilgiler.get("İşlem Referansı", ""))
    ws.cell(row=satir, column=5, value=bilgiler.get("Tutar", ""))

    wb.save(dosya_yolu)

def clean_and_parse_json(dirty_json_string):
    """
    Markdown kod bloğu içine alınmış veya gereksiz boşluklar içeren bir dizeden
    JSON verisini temizler ve ayrıştırır.

    Args:
        dirty_json_string (str): Temizlenmesi gereken JSON dizesi.

    Returns:
        dict: Ayrıştırılmış JSON verisi.
        None: Eğer ayrıştırma başarısız olursa.
    """
    # 1. Baştaki ve sondaki boşlukları (newline dahil) temizle
    # Bu adım, dıştaki tek tırnakların (eğer varsa) ve genel boşlukların temizlenmesine yardımcı olur.
    cleaned_string = dirty_json_string.strip()

    # 2. Markdown kod bloğu işaretlerini kaldır
    # '```json\n' ile başlıyor ve '\n```\n' ile bitiyor varsayıyoruz
    if cleaned_string.startswith('```json\n') and cleaned_string.endswith('\n```'):
        # '```json\n' (8 karakter) ve '\n```' (4 karakter) kaldır
        cleaned_string = cleaned_string[8:-4]
        # Kalan dizedeki baştaki ve sondaki boşlukları tekrar temizle
        cleaned_string = cleaned_string.strip()
    elif cleaned_string.startswith('```\n') and cleaned_string.endswith('\n```'):
        # Eğer sadece '```\n' ile başlıyorsa (4 karakter) ve '\n```' ile bitiyorsa
        cleaned_string = cleaned_string[4:-4]
        cleaned_string = cleaned_string.strip()
    
    # 3. Temizlenmiş dizeyi JSON olarak ayrıştırmaya çalış
    try:
        json_data = json.loads(cleaned_string)
        return json_data
    except json.JSONDecodeError as e:
        print(f"HATA: JSON ayrıştırma başarısız oldu. Hata: {e}")
        print(f"Ayrıştırılamayan dize (temizlendikten sonra): '{cleaned_string}'")
        return None
    except Exception as e:
        print(f"Beklenmedik bir hata oluştu: {e}")
        return None


def query_gemini(prompt, model="gemini-1.5-flash-latest"):
    """
    Belirtilen Gemini modeliyle bir metin sorgusu gönderir.

    Args:
        prompt (str): Modele gönderilecek kullanıcı istemi (prompt).
        model (str): Kullanılacak Gemini modelinin adı. Varsayılan olarak "gemini-1.5-flash-latest".

    Returns:
        str: Modelden gelen yanıtın metin içeriği.
    """
    # API anahtarınızı ortam değişkeninden alın.
    # Güvenlik nedeniyle API anahtarınızı doğrudan kodunuza yazmaktan kaçının.
    # Ortam değişkenini ayarlamak için (örneğin Linux/macOS'ta):
    # export GOOGLE_API_KEY='your_api_key_here'
    # Windows'ta:
    # set GOOGLE_API_KEY='your_api_key_here'

    api_key = config["gemini_api_key"]

    if not api_key:
        print("HATA: GOOGLE_API_KEY ortam değişkeni ayarlanmadı.")
        print("Lütfen API anahtarınızı ayarlayın ve tekrar deneyin.")
        return "API anahtarı eksik."

    # Gemini API'yi API anahtarı ile yapılandırın
    genai.configure(api_key=api_key)

    try:
        # Belirtilen modeli yükleyin
        gemini_model = genai.GenerativeModel(model)

        # Modele istemi gönderin
        response = gemini_model.generate_content(prompt)

        # Yanıtın metin içeriğini döndürün
        return response.text
    
    except Exception as e:
        print(f"Gemini API'ye istek gönderirken bir hata oluştu: {e}")
        return f"Hata: {e}"


def parse_gemini_json_response(response_text):
    try:
        # Gemini'den gelen string JSON formatında olacak
        bilgiler = json.loads(response_text)
        return bilgiler
    except json.JSONDecodeError as e:
        print("JSON decode hatası:", e)
        return {}

def get_subject(msg):
    subject, encoding = decode_header(msg["Subject"])[0]
    if isinstance(subject, bytes):
        return subject.decode(encoding or "utf-8", errors="ignore")
    return subject

def get_body(msg):
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                return part.get_payload(decode=True).decode(errors="ignore")
    else:
        return msg.get_payload(decode=True).decode(errors="ignore")
    return ""

def check_mail():

    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL, PASSWORD)
    mail.select("inbox")
    status, messages = mail.search(None, '(UNSEEN)')
    
    if status != 'OK':
        print("Mail arama başarısız!")
        return
    
    mail_ids = messages[0].split()

    for mail_id in mail_ids:
        _, msg_data = mail.fetch(mail_id, "(RFC822)")
        msg = email.message_from_bytes(msg_data[0][1])
        subject = get_subject(msg)

        if "Dekont" in subject:
            try:
                print("\n--- Yeni Dekont Maili ---")

                # Maildeki PDF ekini bul ve diske kaydet
                pdf_path = None
                for part in msg.walk():
                    content_disposition = part.get("Content-Disposition", "")
                    if "attachment" in content_disposition:
                        filename = part.get_filename()
                        if filename and filename.lower().endswith(".pdf"):
                            # Geçici dosya ismi
                            pdf_path = os.path.join("temp", filename)
                            os.makedirs("temp", exist_ok=True)
                            with open(pdf_path, "wb") as f:
                                f.write(part.get_payload(decode=True))
                            print(f"PDF dosyası kaydedildi: {pdf_path}")
                            break

                if not pdf_path:
                    print("PDF eki bulunamadı.")
                    mail.store(mail_id, '-FLAGS', '\\Seen')  # okunmamış bırak
                    continue

                pdf_text = extract_pdfminer(pdf_path)
                
                prompt = f"""
Aşağıdaki dekont içeriğinden belirtilen bilgileri çıkararak, **yalnızca** ve **kesinlikle** aşağıdaki JSON formatında bir çıktı üretin. Başka hiçbir açıklama, metin veya ek bilgi eklemeyin.

**Çıkarılacak Bilgiler ve Kurallar:**
- "İşlem Tarihi": Dekonttaki işlem tarihini **saat bilgisiyle birlikte** alın.
- "Gönderen Kişi": Dekonttaki gönderen kişinin adını alın.
- "Sorgu Numarası": Dekonttaki sorgu numarasını alın.
- "İşlem Referansı": Dekonttaki işlem referansını alın.
- "Tutar": Dekonttaki "İşlem Tutarı" veya "Tutar" alanlarından hangisi varsa onu alın.

**JSON Çıktı Formatı:**
```json
{{
    "İşlem Tarihi": "",
    "Gönderen Kişi": "",
    "Sorgu Numarası": "",
    "İşlem Referansı": "",
    "Tutar": ""
}}
```

**Dekont İçeriği:**
{pdf_text}
"""

                gemini_response = query_gemini(prompt)  # Gemini API çağrısı, string JSON döner
                
                bilgiler = clean_and_parse_json(gemini_response)

                append_pdf_data_to_excel(bilgiler, EXCEL_YOLU)

                mail.store(mail_id, '+FLAGS', '\\Seen')

                # Geçici pdf dosyasını sil
                os.remove(pdf_path)

            except Exception as e:
                print(f"Hata: {e}")
                mail.store(mail_id, '-FLAGS', '\\Seen')  # okunmamış bırak
                if pdf_path and os.path.exists(pdf_path):
                    os.remove(pdf_path)
                raise    
       
    mail.logout()

if __name__ == "__main__":
   
    while True:
        print("Yeni mailler kontrol ediliyor.")
        check_mail()
        print("10 saniye ara")
        time.sleep(10)
